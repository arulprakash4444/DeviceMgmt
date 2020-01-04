import openpyxl
import XLdeco
from datetime import date, datetime

# (Example array) The Array will be got from json2Array.py
TheArray = [['No', 'Device_Name', 'Alias', 'Team'], ['1', 'iphonese', 'v-arupra@microsoft.com', 'officelens'], ['2', 'iphone6', 'v-amvutu@microsoft.com', 'wxp'], ['3', 'iphone7', 'v-sajara@microsoft.com', 'kaizala'], ['4', 'iphone8', 'v-chered@microsoft.com', 'visio']]


def fileNameGen(table):
    
    today = date.today()
    DATE = today.strftime("%d%b%Y")
    now = datetime.now()
    current_time = now.strftime("%H-%M-%S")

    FILENAME = table + " -- " + DATE + " -- " + current_time
    return FILENAME + ".xlsx"


def createXL(filename):
    wb = openpyxl.Workbook()
    wb.save(filename)
    print("Created "+ filename +"!")


# file = "test123.xlsx"


def XLhorizontalPrinter(array, rowNumber):
    for i, element in enumerate(array):
        sheet.cell(row = rowNumber ,column = i+1).value = element


def XLMaker(Array, file):
    wb = openpyxl.load_workbook(file)
    global sheet
    sheet = wb.active

    for i in range(1, len(Array) + 1):
        XLhorizontalPrinter(Array[i - 1], i)

    wb.save(file)


def array2excel(Array, table):
    filename = fileNameGen(table)
    createXL(filename)
    XLMaker(Array, filename)

    XLdeco.itStartsHere(filename, table)



# if __name__ == "__main__":
#     array2excel(array, table)
