# color codes
# ios_header: #34519A
# ios_accent: #BBD7F0

# android_header: #C45911
# android_accent: #FDE3D4

# red: #FC0200
# green: #8ED058

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.styles import colors




def rowPainter(row_Number, color):
    for i in range(1, maxColumn+1):
        fillattrib = PatternFill(start_color=color, end_color=color, fill_type = "solid")
        sheet.cell(row = row_Number, column = i).fill = fillattrib


def rowChanger(row_Number, Fsize, Fname):
    for i in range(1, maxColumn+1):
        fontattrib = Font(size=Fsize, name=Fname)
        sheet.cell(row = row_Number, column = i).font = fontattrib

        medium = Side(border_style="medium", color="000000")
        borderattrib = Border(left=medium , right=medium, top=medium, bottom=medium)
        sheet.cell(row = row_Number, column = i).border = borderattrib 



# making the header bold and applying color to font and bg
def XLheader(color):
    rowPainter(1, color)

    for i in range(1, maxColumn+1):
        fontHeader = Font(color=colors.WHITE, bold=True, name="Times New Roman", size=11)
        sheet.cell(row = 1, column = i).font = fontHeader




def Painter(color):
    for i in range(3, maxRow+1, 2):
        rowPainter(i, color)





def change4All(size = 11, name="Times New Roman"):
    for i in range(1, maxRow+1):
        rowChanger(i, size, name)



def locker(color):
    aliasColNum = 0
    for i in range(1, maxColumn+1):
        if sheet.cell(row=1, column=i).value == "Alias":
            aliasColNum = i

    for j in range(2, maxRow+1):
        if sheet.cell(row=j, column=aliasColNum).value == "locker":

            fillattrib = PatternFill(start_color=color, end_color=color, fill_type = "solid")
            sheet.cell(row = j, column = aliasColNum).fill = fillattrib



def itStartsHere(name, table):
    filename = name

    ios_header = "34519A"
    ios_accent = "BBD7F0"

    android_header = "C45911"
    android_accent = "FDE3D4"

    # red = "FC0200"
    green = "8ED058"

    if table == "ios":
        header = ios_header
        accent = ios_accent

    elif table == "android":
        header = android_header
        accent = android_accent


    wb = openpyxl.load_workbook(filename)
    global sheet
    sheet = wb.active

    global maxRow
    maxRow = sheet.max_row
    global maxColumn
    maxColumn = sheet.max_column


    Painter(accent)
    change4All()
    XLheader(header)
    locker(green)

    wb.save(filename)



# if __name__ == "__main__":
#     itStartsHere(filename, tabletype)
