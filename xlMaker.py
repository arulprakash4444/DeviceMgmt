import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, PatternFill

# creating a excel file
# file = "testXL.xlsx"
#
# wb = openpyxl.Workbook()
# wb.save(file)

wb = openpyxl.load_workbook("testXL.xlsx")
# wb.create_sheet("sheet_two", 0)

sheet = wb["Sheet"]
for i in range(1, 38, 2):
    for j in range(1, 10):
        fia1 = PatternFill(fill_type="solid", bgColor="FFc7ce")
        fa1 = Font(color=colors.WHITE, name="Times New Roman", size=11)
        sheet.cell(row=i, column=j).value = "prakash"
        sheet.cell(row=i, column=j).font = fa1
        sheet.cell(row=i, column=j).fill = fia1


# print(sheet["A1"].value)
# sheet["a1"].value = "prakash"
# a1 = sheet["A1"]
# fa1 = Font(color=colors.WHITE, name="Times New Roman", size=11)
# a1.font = fa1
# fia1 = PatternFill(fill_type="solid", bgColor="FFc7ce")
# a1.fill = fia1
wb.save("testXL.xlsx")