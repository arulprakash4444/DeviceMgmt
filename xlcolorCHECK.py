import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side

filename = "test123.xlsx"
wb = openpyxl.load_workbook(filename)


sheet = wb.active

# fillattr = PatternFill(fill_type="solid", bgColor=col)
# sheet.cell(row=2, column=6).fill = fillattr
# sheet['A1'].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
# sheet['A1'].fill = PatternFill(start_color=col, end_color=col, fill_type = "solid")
# wb.save(filename)

# for i in range(1, maxColumn+1):
# fillHeader = PatternFill(start_color=col, end_color=col, fill_type = "solid")
# sheet.cell(row = 1, column = 1).fill = fillHeader

# fontHeader = Font(color="FFFFFF",size=11)
# sheet.cell(row = 1, column = 1).font = fontHeader

# wb.save(filename)

medium = Side(border_style="medium", color="000000")
sheet.cell(row=3, column = 2).border = Border(left=medium , right=medium, top=medium, bottom=medium)

wb.save(filename)